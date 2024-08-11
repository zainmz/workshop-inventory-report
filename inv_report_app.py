import os
import glob
import tempfile

import requests
from datetime import datetime
from bs4 import BeautifulSoup
import pandas as pd
import streamlit as st

from openpyxl import Workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils.dataframe import dataframe_to_rows


def get_latest_files(input_folder='inputs'):
    # Get the list of files in the folder
    files = glob.glob(os.path.join(input_folder, '*'))

    # Filter files to include only those with '.xls' or '.htm' extension
    valid_files = [file for file in files if file.endswith('.xls') or file.endswith('.htm')]

    # Check if there are any files found
    if not valid_files:
        raise FileNotFoundError("No valid files found in the directory.")

    # Get the latest file based on modification time
    latest_files = sorted(valid_files, key=os.path.getmtime, reverse=True)

    qty_file = None
    amount_file = None

    for file in latest_files:
        if qty_file is None or amount_file is None:
            soup = parse_html_to_soup(file)
            if 'Opening balance' in soup.text and amount_file is None:
                amount_file = file
            elif 'Opening Stock' in soup.text and qty_file is None:
                qty_file = file
        else:
            break

    return qty_file, amount_file


def save_uploaded_file(uploadedfile, new_extension=None):
    """ Save the uploaded file temporarily and return the file path.
        If new_extension is provided, rename the file with that extension.
    """
    with tempfile.NamedTemporaryFile(delete=False, suffix=new_extension if new_extension else '') as tmp_file:
        tmp_file.write(uploadedfile.getbuffer())
        return tmp_file.name


def parse_html_to_soup(html_file_path):
    # Read the HTML file
    with open(html_file_path, 'r', encoding='utf-8') as file:
        html_content = file.read()

    # Parse the HTML content using BeautifulSoup
    soup = BeautifulSoup(html_content, 'html.parser')

    return soup


def soup_to_dataframe(soup, is_amount_file=False):
    # Find the table in the parsed HTML
    table = soup.find('table', {'id': 'TableResult'})

    # Initialize lists to store data
    items = []
    warehouses = []
    data_rows = []

    # Process the table rows
    for row in table.find_all('tr'):
        # Identify item rows
        if row.find('td') and row.find('td').get('colspan') == '15':
            current_item = row.get_text(strip=True).replace('ITEM : ', '')

        # Process warehouse data rows
        elif row.find('td') and not row.find('td').get('colspan'):
            warehouse_name = row.find_all('td')[0].get_text(strip=True)
            data = [td.get_text(strip=True) for td in row.find_all('td')]

            items.append(current_item)
            warehouses.append(warehouse_name)
            data_rows.append(data)

    if is_amount_file:
        columns = [
            'Warehouse Name', 'Opening Balance', 'Receipt', 'Issue',
            'Transfer Out', 'Transfer In', 'Adjustment (+)', 'Adjustment (-)',
            'Disposal', 'Purchase Return', 'Sales', 'Reserved',
            'Sales Return', 'Damaged', 'Closing Balance'
        ]
    else:
        columns = [
            'Warehouse Name', 'Opening Stock', 'Receipt', 'Issue',
            'Transfer Out', 'Transfer In', 'Adjustment (+)', 'Adjustment (-)',
            'Disposal', 'Purchase Return', 'Sales', 'Reserved',
            'Sales Return', 'Damaged', 'Closing Stock'
        ]

    # Create DataFrame
    df = pd.DataFrame(data_rows, columns=columns)

    # Add the item column to the DataFrame
    df['Item'] = items
    df['Warehouse Name'] = warehouses

    return df


def get_unit_price(item_code):
    """
    Fetch the unit price using the provided API.

    Parameters:
        item_code (str): The code of the item to fetch the price for.

    Returns:
        float: The unit price of the item, or 0 if not found.
    """
    # Extract the SubminorCode (e.g., '6_2' from '6_2_1')
    subminor_code = "_".join(item_code.split("_")[:2])

    # Define the API endpoint
    api_url = "https://tasleeh-ims.com:8023/api/apis/GeItemSegregationBySubMinorCode"

    # Set up the parameters for the request
    params = {'SubminorCode': subminor_code}

    try:
        # Make the GET request to the API
        response = requests.get(api_url, params=params)
        response.raise_for_status()  # Raise an exception for HTTP errors

        # Parse the JSON response
        data = response.json()

        # Check if the API returned a success status
        if data['Status'] == 'Success':
            # Iterate through the items in the response data
            for item in data['Data']:
                if item['ItemCode'] == item_code:
                    return item['BaseUOMUnitPrice']  # Return the desired price

        # If the item was not found, return 0
        return 0.0

    except requests.exceptions.RequestException as e:
        print(f"Error fetching unit price for {item_code}: {e}")
        return 0.0


def add_unit_price_column(df):
    """
    Add a unit price column to the DataFrame using item codes.

    Parameters:
        df (pd.DataFrame): The DataFrame to add the unit price to.

    Returns:
        pd.DataFrame: The DataFrame with the added unit price column.
    """
    df = df.copy()
    df.loc[:, 'Item Code'] = df['Item'].apply(lambda x: x.split(' - ')[0])
    df.loc[:, 'Unit Price'] = df['Item Code'].apply(get_unit_price)
    return df


def merge_dataframes(qty_df, amount_df):
    # Merge dataframes based on the 'Item' and 'Warehouse Name'
    merged_df = pd.merge(qty_df, amount_df, on=['Item', 'Warehouse Name'], suffixes=('_qty', '_amount'))

    # Rename columns to fit the final report structure
    final_columns = {
        'Opening Stock_qty': 'Opening Stock',
        'Unit Price': 'Unit Price',
        'Opening Balance': 'Opening Balance Amount',
        'Transfer In_qty': 'Purchase Stock',
        'Transfer In_amount': 'Total Purchase',
        'Sales_qty': 'Sales Stock',
        'Sales_amount': 'Total Sales',
        'Closing Stock_qty': 'Closing Stock',
        'Closing Balance': 'Closing Balance Amount'
    }

    merged_df = merged_df.rename(columns=final_columns)

    # Reorder the columns to match the final report structure
    organized_df = merged_df[
        [
            'Item', 'Opening Stock', 'Unit Price', 'Opening Balance Amount',
            'Purchase Stock', 'Unit Price', 'Total Purchase',
            'Sales Stock', 'Unit Price', 'Total Sales',
            'Closing Stock', 'Closing Balance Amount'
        ]
    ]

    return organized_df


def apply_excel_formatting(df, output_file):
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Add the data to the worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Merge cells for the headers
    ws.merge_cells('B1:D1')
    ws['B1'] = 'OPENING'

    ws.merge_cells('E1:G1')
    ws['E1'] = 'PURCHASE'

    ws.merge_cells('H1:J1')
    ws['H1'] = 'SALES'

    ws.merge_cells('K1:L1')
    ws['K1'] = 'CLOSING STOCK'

    # Apply styles for headers
    fill_opening = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    fill_purchase = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    fill_sales = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    fill_closing = PatternFill(start_color="EAD1DC", end_color="EAD1DC", fill_type="solid")

    bold_font = Font(bold=True)  # Create a bold font

    # Apply fills and bold font to merged cells
    ws['B1'].fill = fill_opening
    ws['B1'].font = bold_font
    ws['E1'].fill = fill_purchase
    ws['E1'].font = bold_font
    ws['H1'].fill = fill_sales
    ws['H1'].font = bold_font
    ws['K1'].fill = fill_closing
    ws['K1'].font = bold_font

    # Apply alignment to the merged cells
    alignment = Alignment(horizontal="center", vertical="center")
    ws['B1'].alignment = alignment
    ws['E1'].alignment = alignment
    ws['H1'].alignment = alignment
    ws['K1'].alignment = alignment

    # Apply cell colors to corresponding columns
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=4):
        for cell in row:
            cell.fill = fill_opening

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=7):
        for cell in row:
            cell.fill = fill_purchase

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=8, max_col=10):
        for cell in row:
            cell.fill = fill_sales

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=11, max_col=12):
        for cell in row:
            cell.fill = fill_closing

    # Define border style
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Apply border to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Auto-adjust column widths based on the maximum length of data in each column
    for col in ws.columns:
        max_length = 0
        column_letter = None

        for cell in col:
            if not isinstance(cell, MergedCell):
                if column_letter is None:
                    column_letter = cell.column_letter  # Get the column letter from the first non-merged cell

                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

        if column_letter:
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Save the formatted workbook
    wb.save(output_file)


def validate_file_content(soup, expected_keywords):
    """ Validate the content of the HTML file by checking for expected keywords """
    text_content = soup.get_text(strip=True)
    return all(keyword in text_content for keyword in expected_keywords)

def main():
    st.title("Automated Report Generator")

    # Step 1: File upload
    st.header("Upload your HTML or Excel files")
    qty_file = st.file_uploader("Upload the Quantity File", type=["xls", "htm"])
    amount_file = st.file_uploader("Upload the Amount File", type=["xls", "htm"])

    if st.button("Generate Report"):
        if qty_file and amount_file:
            # Initialize the progress bar
            progress_bar = st.progress(0)
            progress = 0

            # Handle renaming if an .xls file is uploaded
            if qty_file.name.endswith('.xls'):
                qty_file_path = save_uploaded_file(qty_file, new_extension=".htm")
            else:
                qty_file_path = save_uploaded_file(qty_file)

            if amount_file.name.endswith('.xls'):
                amount_file_path = save_uploaded_file(amount_file, new_extension=".htm")
            else:
                amount_file_path = save_uploaded_file(amount_file)

            progress += 10
            progress_bar.progress(progress)

            # Parse the files
            qty_soup = parse_html_to_soup(qty_file_path)
            amount_soup = parse_html_to_soup(amount_file_path)

            # Validate the file contents
            if not validate_file_content(qty_soup, ["Opening Stock"]):
                st.error("The Quantity file does not appear to be correct. Please upload the correct Quantity file.")
                return

            if not validate_file_content(amount_soup, ["Opening balance"]):
                st.error("The Amount file does not appear to be correct. Please upload the correct Amount file.")
                return

            progress += 20
            progress_bar.progress(progress)

            # Convert the parsed HTML to DataFrames
            qty_df = soup_to_dataframe(qty_soup, is_amount_file=False)
            amount_df = soup_to_dataframe(amount_soup, is_amount_file=True)

            progress += 20
            progress_bar.progress(progress)

            # Filter for 'WS1 - WS Shuwaikh' rows only
            qty_df = qty_df[qty_df['Warehouse Name'].str.contains('WS1 - WS Shuwaikh', na=False)]
            amount_df = amount_df[amount_df['Warehouse Name'].str.contains('WS1 - WS Shuwaikh', na=False)]

            # Add Unit Price column
            qty_df = add_unit_price_column(qty_df)

            progress += 20
            progress_bar.progress(progress)

            # Merge the DataFrames
            final_report_df = merge_dataframes(qty_df, amount_df)

            progress += 10
            progress_bar.progress(progress)

            # Apply Excel formatting and save the final report
            output_file = "Final_Report_Structured.xlsx"
            apply_excel_formatting(final_report_df, output_file)

            progress += 10
            progress_bar.progress(progress)

            st.success(f"Report generated and saved as {output_file}")

            # Provide a download link for the file
            with open(output_file, "rb") as file:
                st.download_button(
                    label="Download the report",
                    data=file,
                    file_name=output_file,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.error("Please upload both the Quantity and Amount files.")

if __name__ == "__main__":
    main()
